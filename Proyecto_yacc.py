
# PARTE 2: ANALISIS GRAMATICAL (PARSING) USANDO YACC
# DE LA ENTRADA YA TOKENIZADA CON LEX 
from openpyxl import Workbook
import json
import ply.yacc as yacc
from Proyecto_lex import tokens
fila=2
fila2=2
fila3=2
global_titulo='Star Wars Episode IV - A New Hope'
def p_start(p):
    's : KEY_S PELICULAS DOSPUNTOS CORCHETE_I a CORCHETE_F KEY_F'
    p[0] = p[1] + p[2] + p[3]+ p[4] + p[5]+ p[6] +p[7]
  
#a es lo que esta adentro de los corchetes principales de las peliculas
def p_datos(p):
    'a :  KEY_S p KEY_F b'
    p[0] = p[1] + p[2] + p[3] +p[4]
   
    
def p_datos_b(p):
    'b : COMA KEY_S p KEY_F b'
    p[0] = p[1] + p[2] + p[3]+ p[4] +p[5]
  
    
def p_datos_b_1(p):
    'b : '
    p[0] = ''

#No hay películas
def p_epsilon(p):
    'a : '
    p[0]=''
#p es el cuerpo de cada pelicula
def p_peli(p):
    'p : TITULO DOSPUNTOS COMILLAS VALUE COMILLAS COMA ANIO DOSPUNTOS VALUE COMA GENERO DOSPUNTOS COMILLAS VALUE COMILLAS COMA DURACION DOSPUNTOS COMILLAS VALUE COMILLAS COMA DIRECTOR DOSPUNTOS COMILLAS VALUE COMILLAS COMA ESTRELLAS DOSPUNTOS CORCHETE_I array CORCHETE_F COMA CALIIMB DOSPUNTOS VALUE'
    p[0] = p[1] + p[2] + p[3] + p[4] + p[5] + p[6]+ p[7] +p[8] +p[9] + p[10] + p[11]+ p[12] +p[13]+ p[14] + p[15] + p[16]+ p[17] +p[18]+  p[19] +p[20]+p[21]+ p[22] +p[23]+ p[24] + p[25] + p[26]+ p[27] +p[28] +  p[29] +p[30]+p[31]+ p[32] +p[33]+ p[34] + p[35] + p[36]+ p[37] 
    global global_titulo
    global fila
    global fila2
    global_titulo=p[4]
    sheet1['A'+str(fila)]=p[4]
    sheet1['B'+str(fila)]=p[9]
    sheet1['C'+str(fila)]=p[14]
    sheet1['D'+str(fila)]=p[20]
    sheet1['E'+str(fila)]=p[37]
    sheet2['A'+str(fila2)]=fila
    sheet2['B'+str(fila2)]=p[26]
    sheet1['F'+str(fila)]=fila
    fila+=1
    fila2+=1
def p_peli2(p):
    'p : TITULO DOSPUNTOS COMILLAS VALUE COMILLAS COMA ANIO DOSPUNTOS VALUE COMA GENERO DOSPUNTOS COMILLAS VALUE COMILLAS COMA DURACION DOSPUNTOS COMILLAS VALUE COMILLAS COMA DIRECTORES DOSPUNTOS CORCHETE_I array2 CORCHETE_F COMA ESTRELLAS DOSPUNTOS CORCHETE_I array CORCHETE_F COMA CALIIMB DOSPUNTOS VALUE'
    p[0] = p[1] + p[2] + p[3] + p[4] + p[5] + p[6]+ p[7] +p[8] +p[9] + p[10] + p[11]+ p[12] +p[13]+ p[14] + p[15] + p[16]+ p[17] +p[18]+  p[19] +p[20]+p[21]+ p[22] +p[23]+ p[24] + p[25] + p[26]+ p[27] +p[28] +  p[29] +p[30]+p[31]+ p[32] +p[33]+ p[34] + p[35] + p[36]+ p[37] 
    global global_titulo
    global fila
    global_titulo=p[4]
    sheet1['A'+str(fila)]=p[4]
    sheet1['B'+str(fila)]=p[9]
    sheet1['C'+str(fila)]=p[14]
    sheet1['D'+str(fila)]=p[20]
    sheet1['E'+str(fila)]=p[37]
    sheet1['F'+str(fila)]=fila
    fila+=1
    
def p_array(p):
    'array : COMILLAS VALUE COMILLAS COMA COMILLAS VALUE COMILLAS array1'
    p[0] = p[1] + p[2] + p[3]+ p[4] +p[5] +p[6]+p[7] +p[8]
    global fila3
    sheet3['A'+str(fila3)]=fila
    sheet3['B'+str(fila3)]=p[2]
    fila3+=1
    sheet3['A'+str(fila3)]=fila
    sheet3['B'+str(fila3)]=p[6]
    fila3+=1
    
def p_array1(p):
    'array1 : COMA COMILLAS VALUE COMILLAS array1 '
    p[0] = p[1] + p[2] + p[3]+ p[4] +p[5]
    global fila3
    sheet3['A'+str(fila3)]=fila
    sheet3['B'+str(fila3)]=p[3]
    fila3+=1
    
def p_array1e(p):
    'array1 : '
    p[0] = ''
    print(p[0])

def p_array2(p):
    'array2 : COMILLAS VALUE COMILLAS COMA COMILLAS VALUE COMILLAS array3'
    p[0] = p[1] + p[2] + p[3]+ p[4] +p[5] +p[6]+p[7] +p[8]
    global fila2
    sheet2['A'+str(fila2)]=fila
    sheet2['B'+str(fila2)]=p[2]
    fila2+=1
    sheet2['A'+str(fila2)]=fila
    sheet2['B'+str(fila2)]=p[6]
    fila2+=1
    
def p_array12(p):
    'array3 : COMA COMILLAS VALUE COMILLAS array3 '
    p[0] = p[1] + p[2] + p[3]+ p[4] +p[5]
    global fila2
    sheet2['A'+str(fila2)]=fila
    sheet2['B'+str(fila2)]=p[3]
    fila2+=1
    
def p_array1e2(p):
    'array3 : '
    p[0] = ''
   

"""#no sabemos bien de esto
def p_body_array(p):
    'array : VALUE COMA VALUE array '
    p[0] = p[1] + p[2] + p[3]+ p[4]"""
    
# Error rule for syntax errors
def p_error(p):
    print("Syntax error in input!")


# Build the parser
parser = yacc.yacc()


with open('archivos.json', 'r', encoding='utf-8') as file:
    try:
        json_data = json.load(file)
        input_str = json.dumps(json_data,ensure_ascii=False)
        workbook = Workbook()
        sheet1 = workbook.active  # Hoja activa por defecto (primera hoja)
        sheet1.title = 'Peliculas'   # Renombrar la hoja por defecto   
        sheet2 = workbook.create_sheet(title='Directores')
        sheet3 = workbook.create_sheet(title='Estrellas')
        sheet1['A1']="Titulo"
        sheet1['B1']="Año"
        sheet1['C1']="Genero"
        sheet1['D1']="Duración"
        sheet1['E1']="Calificación"
        sheet1['F1']="ID"
        sheet2['A1']="Id"
        sheet2['B1']="Nombre"
        sheet3['A1']="Id"
        sheet3['B1']="Nombre"
        result = parser.parse(input_str)      
        print( result)
        workbook.save('./mi_archivo.xlsx')
    except json.decoder.JSONDecodeError as e:
        print(f"Error de sintaxis en el archivo JSON: {e}")
    except Exception as e:
        print(f"Error: {e}")
